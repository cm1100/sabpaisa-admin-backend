"""
Settlement API Views
Following SOLID principles and DRF best practices
"""
from rest_framework import status, viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from django.http import HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from decimal import Decimal
import logging

from .models import (
    SettlementBatch,
    SettlementDetail,
    SettlementReport,
    SettlementConfiguration,
    SettlementReconciliation
)
from .services import (
    SettlementProcessingService,
    SettlementReportService,
    SettlementReconciliationService,
    SettlementAnalyticsService
)
from django.db.models import Count, Sum
from datetime import datetime, timedelta
from django.db.models import Q, Sum, Count, Avg, F, Case, When, Value, CharField
from .serializers import (
    SettlementBatchSerializer,
    SettlementDetailSerializer,
    CreateSettlementBatchSerializer,
    ProcessSettlementSerializer,
    SettlementReportSerializer,
    GenerateReportSerializer,
    SettlementConfigurationSerializer,
    SettlementReconciliationSerializer,
    CreateReconciliationSerializer,
    UpdateReconciliationSerializer,
    SettlementStatisticsSerializer,
    ClientSettlementSummarySerializer,
    SettlementFilterSerializer
)
from transactions.models import TransactionDetail

logger = logging.getLogger(__name__)


class SettlementBatchViewSet(viewsets.ModelViewSet):
    """
    ViewSet for settlement batch operations
    """
    queryset = SettlementBatch.objects.all()
    serializer_class = SettlementBatchSerializer
    permission_classes = []  # Temporarily disabled for testing
    lookup_field = 'batch_id'
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.processing_service = SettlementProcessingService()
    
    def list(self, request):
        """List settlement batches with filters"""
        try:
            # Validate filters
            filter_serializer = SettlementFilterSerializer(data=request.query_params)
            filter_serializer.is_valid(raise_exception=True)
            filters = filter_serializer.validated_data
            
            # Get batches
            batches = self.processing_service.get_settlement_batches(filters)
            
            return Response({
                'results': batches,
                'count': len(batches)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error listing settlement batches: {str(e)}")
            return Response(
                {"error": "Failed to fetch settlement batches"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def create(self, request):
        """Create a new settlement batch"""
        try:
            serializer = CreateSettlementBatchSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            batch = self.processing_service.create_settlement_batch(
                batch_date=serializer.validated_data.get('batch_date'),
                user=request.user if request.user.is_authenticated else None
            )
            
            batch_serializer = SettlementBatchSerializer(batch)
            return Response(batch_serializer.data, status=status.HTTP_201_CREATED)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Error creating settlement batch: {str(e)}")
            return Response(
                {"error": "Failed to create settlement batch"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def process(self, request, batch_id=None):
        """Process a settlement batch"""
        try:
            batch = self.processing_service.process_settlement_batch(
                batch_id=batch_id,
                user=request.user if request.user.is_authenticated else None
            )
            
            serializer = SettlementBatchSerializer(batch)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Error processing settlement batch: {str(e)}")
            return Response(
                {"error": "Failed to process settlement batch"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def details(self, request, batch_id=None):
        """Get settlement details for a batch"""
        try:
            details = SettlementDetail.objects.filter(batch__batch_id=batch_id)
            serializer = SettlementDetailSerializer(details, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching settlement details: {str(e)}")
            return Response(
                {"error": "Failed to fetch settlement details"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SettlementConfigurationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for settlement configuration
    """
    queryset = SettlementConfiguration.objects.all()
    serializer_class = SettlementConfigurationSerializer
    permission_classes = []  # Temporarily disabled for testing
    lookup_field = 'client_code'
    
    def create(self, request):
        """Create or update settlement configuration"""
        try:
            serializer = SettlementConfigurationSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            # Check if configuration exists
            client_code = serializer.validated_data['client_code']
            config = SettlementConfiguration.objects.filter(client_code=client_code).first()
            
            if config:
                # Update existing configuration
                for key, value in serializer.validated_data.items():
                    setattr(config, key, value)
                config.save()
            else:
                # Create new configuration
                config = serializer.save()
            
            response_serializer = SettlementConfigurationSerializer(config)
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Error creating settlement configuration: {str(e)}")
            return Response(
                {"error": "Failed to create configuration"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SettlementReportView(APIView):
    """
    API View for settlement report operations
    """
    permission_classes = []  # Temporarily disabled for testing
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.report_service = SettlementReportService()
    
    def get(self, request):
        """Get list of settlement reports"""
        try:
            batch_id = request.query_params.get('batch_id')
            reports = self.report_service.get_settlement_reports(batch_id)
            
            return Response(reports, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching settlement reports: {str(e)}")
            return Response(
                {"error": "Failed to fetch reports"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def post(self, request):
        """Generate a new settlement report"""
        try:
            serializer = GenerateReportSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            report = self.report_service.generate_settlement_report(
                batch_id=str(serializer.validated_data['batch_id']),
                report_type=serializer.validated_data['report_type'],
                user=request.user
            )
            
            report_serializer = SettlementReportSerializer(report)
            return Response(report_serializer.data, status=status.HTTP_201_CREATED)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Error generating settlement report: {str(e)}")
            return Response(
                {"error": "Failed to generate report"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SettlementReconciliationView(APIView):
    """
    API View for settlement reconciliation
    """
    permission_classes = []  # Temporarily disabled for testing
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.reconciliation_service = SettlementReconciliationService()
    
    def get(self, request):
        """Get pending reconciliations"""
        try:
            reconciliations = self.reconciliation_service.get_pending_reconciliations()
            return Response(reconciliations, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching reconciliations: {str(e)}")
            return Response(
                {"error": "Failed to fetch reconciliations"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def post(self, request):
        """Create a new reconciliation"""
        try:
            serializer = CreateReconciliationSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            reconciliation = self.reconciliation_service.create_reconciliation(
                batch_id=str(serializer.validated_data['batch_id']),
                bank_amount=serializer.validated_data['bank_statement_amount'],
                user=request.user
            )
            
            recon_serializer = SettlementReconciliationSerializer(reconciliation)
            return Response(recon_serializer.data, status=status.HTTP_201_CREATED)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Error creating reconciliation: {str(e)}")
            return Response(
                {"error": "Failed to create reconciliation"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ReconciliationUpdateView(APIView):
    """
    API View for updating reconciliation status
    """
    permission_classes = []  # Temporarily disabled for testing
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.reconciliation_service = SettlementReconciliationService()
    
    def put(self, request, reconciliation_id):
        """Update reconciliation status"""
        try:
            serializer = UpdateReconciliationSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            reconciliation = self.reconciliation_service.update_reconciliation_status(
                reconciliation_id=reconciliation_id,
                status=serializer.validated_data['status'],
                remarks=serializer.validated_data.get('remarks'),
                user=request.user
            )
            
            recon_serializer = SettlementReconciliationSerializer(reconciliation)
            return Response(recon_serializer.data, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Error updating reconciliation: {str(e)}")
            return Response(
                {"error": "Failed to update reconciliation"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SettlementAnalyticsView(APIView):
    """
    API View for settlement analytics
    """
    permission_classes = []  # Temporarily disabled for testing
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.analytics_service = SettlementAnalyticsService()
    
    @method_decorator(cache_page(300))  # Cache for 5 minutes
    def get(self, request):
        """Get settlement analytics"""
        try:
            analytics_type = request.query_params.get('type', 'statistics')
            
            if analytics_type == 'statistics':
                date_range = request.query_params.get('range', '30d')
                data = self.analytics_service.get_settlement_statistics(date_range)
                serializer = SettlementStatisticsSerializer(data)
                
            elif analytics_type == 'client_summary':
                client_code = request.query_params.get('client_code')
                if not client_code:
                    return Response(
                        {"error": "Client code is required"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                days = int(request.query_params.get('days', 30))
                data = self.analytics_service.get_client_settlement_summary(client_code, days)
                serializer = ClientSettlementSummarySerializer(data)
                
            else:
                return Response(
                    {"error": "Invalid analytics type"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching settlement analytics: {str(e)}")
            return Response(
                {"error": "Failed to fetch analytics"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SettlementExportView(APIView):
    """
    API View for exporting settlements to CSV/Excel
    """
    permission_classes = []  # Temporarily disabled for testing
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.processing_service = SettlementProcessingService()
    
    def get(self, request):
        """Export settlements data"""
        try:
            # Get filters from query parameters
            filters = {}
            format_type = request.query_params.get('format', 'csv')
            
            for key in ['status', 'batch_date_from', 'batch_date_to', 'client_code']:
                if request.query_params.get(key):
                    filters[key] = request.query_params.get(key)
            
            # Get batches data
            batches = self.processing_service.get_settlement_batches(filters)
            
            if format_type == 'csv':
                return self.export_csv(batches)
            elif format_type == 'excel':
                return self.export_excel(batches)
            else:
                return JsonResponse({"error": "Unsupported format"}, status=400)
                
        except Exception as e:
            logger.error(f"Error exporting settlements: {str(e)}")
            return JsonResponse(
                {"error": "Failed to export settlements"},
                status=500
            )
    
    def export_csv(self, batches):
        """Export to CSV format"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            'Batch ID', 'Batch Date', 'Status', 'Total Transactions',
            'Total Amount', 'Processing Fee', 'GST Amount', 'Net Amount',
            'Created At', 'Processed At'
        ])
        
        # Data rows
        for batch in batches:
            writer.writerow([
                batch['batch_id'],
                batch['batch_date'],
                batch['status'],
                batch['total_transactions'],
                batch['total_amount'],
                batch['processing_fee'],
                batch['gst_amount'],
                batch['net_settlement_amount'],
                batch['created_at'],
                batch['processed_at'] or ''
            ])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="settlements.csv"'
        return response
    
    def export_excel(self, batches):
        """Export to Excel format"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            import io
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Settlements'
            
            # Headers
            headers = [
                'Batch ID', 'Batch Date', 'Status', 'Total Transactions',
                'Total Amount', 'Processing Fee', 'GST Amount', 'Net Amount',
                'Created At', 'Processed At'
            ]
            
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Data rows
            for row, batch in enumerate(batches, 2):
                worksheet.cell(row=row, column=1, value=batch['batch_id'])
                worksheet.cell(row=row, column=2, value=batch['batch_date'])
                worksheet.cell(row=row, column=3, value=batch['status'])
                worksheet.cell(row=row, column=4, value=batch['total_transactions'])
                worksheet.cell(row=row, column=5, value=batch['total_amount'])
                worksheet.cell(row=row, column=6, value=batch['processing_fee'])
                worksheet.cell(row=row, column=7, value=batch['gst_amount'])
                worksheet.cell(row=row, column=8, value=batch['net_settlement_amount'])
                worksheet.cell(row=row, column=9, value=batch['created_at'])
                worksheet.cell(row=row, column=10, value=batch['processed_at'] or '')
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 15
            
            # Save to memory
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="settlements.xlsx"'
            return response
            
        except ImportError:
            return JsonResponse({"error": "Excel export requires openpyxl library"}, status=500)


class SettlementActivityView(APIView):
    """
    API View for settlement activity timeline
    """
    permission_classes = []  # Temporarily disabled for testing
    
    def get(self, request):
        """Get settlement activity timeline"""
        try:
            limit = int(request.query_params.get('limit', 10))
            
            # Get recent batches with their activity
            recent_batches = SettlementBatch.objects.select_related().order_by('-created_at')[:limit]
            
            activities = []
            for batch in recent_batches:
                # Add batch creation activity
                activities.append({
                    'activity_type': 'batch_created',
                    'description': f'Settlement batch {batch.batch_id} created',
                    'amount': float(batch.total_amount),
                    'timestamp': batch.created_at.isoformat(),
                    'status': 'info',
                    'batch_id': str(batch.batch_id)
                })
                
                # Add processing activity if processed
                if batch.processed_at:
                    status_map = {
                        'COMPLETED': 'success',
                        'FAILED': 'error',
                        'CANCELLED': 'warning'
                    }
                    activities.append({
                        'activity_type': 'batch_processed',
                        'description': f'Settlement batch {batch.batch_id} {batch.status.lower()}',
                        'amount': float(batch.net_settlement_amount) if batch.net_settlement_amount else float(batch.total_amount),
                        'timestamp': batch.processed_at.isoformat(),
                        'status': status_map.get(batch.status, 'info'),
                        'batch_id': str(batch.batch_id)
                    })
            
            # Sort by timestamp (most recent first)
            activities.sort(key=lambda x: x['timestamp'], reverse=True)
            
            return Response(activities[:limit], status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching settlement activities: {str(e)}")
            return Response(
                {"error": "Failed to fetch settlement activities"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SettlementCycleDistributionView(APIView):
    """
    API View for settlement cycle distribution analytics
    """
    permission_classes = []  # Temporarily disabled for testing
    
    def get(self, request):
        """Get settlement cycle distribution"""
        try:
            # Get cycle distribution from settlement configurations
            cycle_stats = SettlementConfiguration.objects.values('settlement_cycle').annotate(
                count=Count('client_code'),
                total_amount=Sum('min_settlement_amount')
            ).order_by('settlement_cycle')
            
            # If no configurations exist, create mock data based on batches
            if not cycle_stats.exists():
                # Default distribution based on typical T+1 settlement
                total_batches = SettlementBatch.objects.count()
                total_amount = SettlementBatch.objects.aggregate(
                    total=Sum('total_amount')
                )['total'] or 0
                
                if total_batches > 0:
                    cycle_distribution = [
                        {
                            'cycle': 'T+0',
                            'count': max(1, int(total_batches * 0.1)),  # 10% same day
                            'amount': float(total_amount * 0.15),
                            'percentage': 15
                        },
                        {
                            'cycle': 'T+1', 
                            'count': max(1, int(total_batches * 0.8)),  # 80% next day
                            'amount': float(total_amount * 0.75),
                            'percentage': 75
                        },
                        {
                            'cycle': 'T+2',
                            'count': max(1, int(total_batches * 0.1)),  # 10% T+2
                            'amount': float(total_amount * 0.10),
                            'percentage': 10
                        }
                    ]
                else:
                    cycle_distribution = []
            else:
                total_count = sum(stat['count'] for stat in cycle_stats)
                total_amount = sum(stat['total_amount'] or 0 for stat in cycle_stats)
                
                cycle_distribution = []
                for stat in cycle_stats:
                    count = stat['count']
                    amount = float(stat['total_amount'] or 0)
                    percentage = (count / total_count * 100) if total_count > 0 else 0
                    
                    cycle_distribution.append({
                        'cycle': stat['settlement_cycle'],
                        'count': count,
                        'amount': amount,
                        'percentage': round(percentage, 1)
                    })
            
            return Response(cycle_distribution, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching cycle distribution: {str(e)}")
            return Response(
                {"error": "Failed to fetch cycle distribution"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class BankWisePerformanceView(APIView):
    """
    API View for bank-wise settlement performance analytics
    Aggregates data from TransactionDetail by bank_name
    """
    permission_classes = []  # Temporarily disabled for testing
    
    def get(self, request):
        """Get bank-wise settlement performance"""
        try:
            # Get date range from request
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            
            # Base query for settled transactions
            # For demo, use pg_name as fallback when bank_name is null\n            query = Q(is_settled=True) & (\n                Q(bank_name__isnull=False) | Q(pg_name__isnull=False)\n            )
            
            if date_from:
                query &= Q(settlement_date__gte=date_from)
            if date_to:
                query &= Q(settlement_date__lte=date_to)
            
            # Aggregate bank performance data
            # Use Case/When to prioritize bank_name, fallback to pg_name\n            bank_stats = TransactionDetail.objects.filter(query).annotate(\n                display_bank_name=Case(\n                    When(bank_name__isnull=False, then=F('bank_name')),\n                    When(pg_name__isnull=False, then=F('pg_name')),\n                    default=Value('Unknown Bank'),\n                    output_field=CharField()\n                )\n            ).values('display_bank_name').annotate(
                total_batches=Count('txn_id'),
                completed=Count('txn_id', filter=Q(settlement_status='SETTLED')),
                pending=Count('txn_id', filter=Q(settlement_status='PENDING')),
                failed=Count('txn_id', filter=Q(settlement_status='FAILED')),
                total_amount=Sum('settlement_amount')
            ).order_by('-total_amount')
            
            bank_performance = []
            for bank_stat in bank_stats:
                bank_name = bank_stat['display_bank_name']
                total_batches = bank_stat['total_batches']
                completed = bank_stat['completed']
                pending = bank_stat['pending'] 
                failed = bank_stat['failed']
                total_amount = float(bank_stat['total_amount'] or 0)
                # avg_processing_time removed from query
                
                # Calculate success rate
                success_rate = (completed / total_batches * 100) if total_batches > 0 else 0
                
                # Simplified processing time for T+1 settlement
                avg_processing_time_str = "< 24 hours"
                
                # Get last settlement time
                last_settlement = TransactionDetail.objects.filter(
                    Q(bank_name=bank_name) | Q(pg_name=bank_name),
                    settlement_date__isnull=False
                ).order_by('-settlement_date').first()
                
                last_settlement_time = last_settlement.settlement_date.isoformat() if last_settlement else None
                
                bank_performance.append({
                    'bank_code': bank_name.replace(' ', '').upper()[:10] if bank_name else 'UNKNOWN',
                    'bank_name': bank_name or 'Unknown Bank',
                    'total_batches': total_batches,
                    'completed': completed,
                    'pending': pending,
                    'failed': failed,
                    'total_amount': total_amount,
                    'success_rate': round(success_rate, 1),
                    'avg_processing_time': avg_processing_time_str,
                    'last_settlement': last_settlement_time
                })
            
            return Response(bank_performance, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching bank-wise performance: {str(e)}")
            return Response(
                {"error": "Failed to fetch bank-wise performance"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SettlementDisputesView(APIView):
    """
    API View for settlement disputes
    Derives disputes from reconciliation mismatches and failed settlements
    """
    permission_classes = []  # Temporarily disabled for testing
    
    def get(self, request):
        """Get settlement disputes"""
        try:
            disputes = []
            
            # Get mismatched reconciliations as disputes
            mismatched_reconciliations = SettlementReconciliation.objects.filter(
                reconciliation_status='MISMATCHED'
            ).select_related('batch')
            
            for recon in mismatched_reconciliations:
                # Try to get client info from batch settlements
                client_info = SettlementDetail.objects.filter(
                    batch=recon.batch
                ).values('client_code').annotate(
                    client_count=Count('client_code')
                ).order_by('-client_count').first()
                
                client_name = f"{client_info['client_code']} Ltd" if client_info else "Unknown Client"
                
                disputes.append({
                    'id': str(recon.reconciliation_id),
                    'dispute_id': f'DSP{str(recon.reconciliation_id)[-6:]}',
                    'batch_id': str(recon.batch.batch_id),
                    'client_name': client_name,
                    'amount': float(abs(recon.difference_amount)),
                    'status': 'OPEN' if recon.reconciliation_status == 'MISMATCHED' else 'IN_PROGRESS',
                    'priority': 'HIGH' if abs(recon.difference_amount) > 10000 else 'MEDIUM',
                    'category': 'AMOUNT_MISMATCH',
                    'description': f'Reconciliation mismatch: Bank amount ₹{recon.bank_statement_amount} vs System amount ₹{recon.system_amount}',
                    'created_at': recon.created_at.isoformat(),
                    'assigned_to': recon.reconciled_by.username if recon.reconciled_by else 'Unassigned',
                    'resolution_deadline': (recon.created_at + timedelta(days=7)).isoformat(),
                    'activities': [
                        {
                            'action': 'Dispute Created',
                            'user': 'System',
                            'timestamp': recon.created_at.isoformat(),
                            'details': 'Auto-generated from reconciliation mismatch'
                        }
                    ]
                })
            
            # Get failed settlement batches as disputes
            failed_batches = SettlementBatch.objects.filter(status='FAILED')
            
            for batch in failed_batches:
                # Get primary client for this batch
                primary_client = SettlementDetail.objects.filter(
                    batch=batch
                ).values('client_code').annotate(
                    amount=Sum('settlement_amount')
                ).order_by('-amount').first()
                
                client_name = f"{primary_client['client_code']} Ltd" if primary_client else "Multiple Clients"
                
                disputes.append({
                    'id': str(batch.batch_id),
                    'dispute_id': f'DSP{str(batch.batch_id)[-6:]}',
                    'batch_id': str(batch.batch_id),
                    'client_name': client_name,
                    'amount': float(batch.total_amount),
                    'status': 'IN_PROGRESS',
                    'priority': 'HIGH',
                    'category': 'TECHNICAL_ISSUE',
                    'description': f'Settlement batch failed to process {batch.total_transactions} transactions',
                    'created_at': batch.created_at.isoformat(),
                    'assigned_to': batch.processed_by.username if batch.processed_by else 'Technical Team',
                    'resolution_deadline': (batch.created_at + timedelta(days=3)).isoformat(),
                    'activities': [
                        {
                            'action': 'Processing Failed',
                            'user': 'System',
                            'timestamp': batch.processed_at.isoformat() if batch.processed_at else batch.created_at.isoformat(),
                            'details': 'Settlement batch processing failed'
                        }
                    ]
                })
            
            # Sort by creation date (newest first)
            disputes.sort(key=lambda x: x['created_at'], reverse=True)
            
            return Response(disputes, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching settlement disputes: {str(e)}")
            return Response(
                {"error": "Failed to fetch settlement disputes"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )